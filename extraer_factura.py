import inspect
import logging
import os
import re
import shutil
import traceback
from collections import defaultdict
from datetime import date, datetime

from dotenv import load_dotenv
from openpyxl import load_workbook

import db
import smtp
import util

load_dotenv()

log_level = os.getenv('LOG_LEVEL', 'INFO')
numeric_level = getattr(logging, log_level.upper(), logging.INFO)

logging.basicConfig(
    level=numeric_level,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)


REQUIRED_COLUMNS = {
    'factura_oficial': ['factura oficial'],
    'fecha_factura': ['fecha facura', 'fecha factura'],
    'remito': ['remito'],
    'numero_serie': ['numero de serie'],
    'total_pedido': ['total pedido'],
    'nombre_material': ['nombre material'],
}


class ExcelRowError(util.PDFInconsistente):
    def __init__(self, message, archivo, hoja=None, fila=None, factura=None):
        super().__init__(message)
        self.archivo = archivo
        self.hoja = hoja
        self.fila = fila
        self.factura = factura

    def describe(self):
        parts = [f'archivo={self.archivo}']
        if self.hoja is not None:
            parts.append(f'hoja={self.hoja}')
        if self.fila is not None:
            parts.append(f'fila={self.fila}')
        if self.factura is not None:
            parts.append(f'factura={self.factura}')
        parts.append(f'detalle={self.args[0]}')
        return ' | '.join(parts)


def normalize_header(value):
    if value is None:
        return ''
    return ' '.join(str(value).strip().lower().split())


def find_required_columns(headers):
    index_by_header = {normalize_header(header): idx for idx, header in enumerate(headers)}
    resolved = {}
    missing = []

    for logical_name, aliases in REQUIRED_COLUMNS.items():
        idx = next((index_by_header[alias] for alias in aliases if alias in index_by_header), None)
        if idx is None:
            missing.append('/'.join(aliases))
        else:
            resolved[logical_name] = idx

    if missing:
        raise util.PDFInconsistente(
            f'No se encontraron columnas obligatorias: {", ".join(missing)}'
        )

    return resolved


def format_excel_date(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if value is None or str(value).strip() == '':
        return None

    raw_value = str(value).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(raw_value, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue

    raise util.PDFInconsistente(f'Fecha de factura inválida: {raw_value}')


def normalize_factura(value):
    if value is None or str(value).strip() == '':
        raise util.PDFInconsistente('La factura oficial está vacía.')

    factura = re.sub(r'\s+', '', str(value).strip())
    if 'C' in factura.upper():
        left, right = re.split(r'[cC]', factura, maxsplit=1)
    elif '-' in factura:
        left, right = factura.split('-', 1)
    else:
        digits = re.sub(r'\D', '', factura)
        if len(digits) < 12:
            raise util.PDFInconsistente(f'Formato de factura inválido: {factura}')
        left, right = digits[:-8], digits[-8:]

    left_digits = re.sub(r'\D', '', left)
    right_digits = re.sub(r'\D', '', right)
    if not left_digits or not right_digits:
        raise util.PDFInconsistente(f'Formato de factura inválido: {factura}')

    return f'{left_digits.zfill(4)}-{right_digits.zfill(8)}'


def extract_nro_remito(value):
    if value is None or str(value).strip() == '':
        raise util.PDFInconsistente('El remito está vacío.')

    remito = str(value).strip().upper()
    match = re.search(r'R(\d+)$', remito)
    if match:
        return int(match.group(1))

    digits = re.findall(r'\d+', remito)
    if digits:
        return int(digits[-1])

    raise util.PDFInconsistente(f'No se pudo extraer el número de remito desde: {value}')


def to_float(value):
    if isinstance(value, (int, float)):
        return float(value)

    if value is None or str(value).strip() == '':
        raise util.PDFInconsistente('Total Pedido vacío.')

    normalized = str(value).strip()
    return float(util.convert_decimal_from_spanish_to_english_format(normalized))


def expand_range_token(token):
    start_text, end_text = [part.strip() for part in token.split('/', 1)]
    start_text = util.normalizar_numero_certificado(start_text)
    end_text = util.normalizar_numero_certificado(end_text)
    if not start_text or not end_text or not start_text.isdigit() or not end_text.isdigit():
        return [token]

    start = int(start_text)
    end = int(end_text)
    if end < start:
        raise util.PDFInconsistente(f'Rango de certificados inválido: {token}')

    width = max(len(start_text), len(end_text))
    return [str(number).zfill(width) for number in range(start, end + 1)]


def parse_certificados(value):
    if value is None or str(value).strip() == '':
        return ['']

    certificados = []
    for token in str(value).split(','):
        token = token.strip()
        if not token:
            continue
        if '/' in token:
            certificados.extend(expand_range_token(token))
        else:
            certificados.append(util.normalizar_numero_certificado(token))

    if not certificados:
        return ['']

    return certificados


def extraer_facturas_desde_xlsx(ruta_xlsx):
    archivo = os.path.basename(ruta_xlsx)
    workbook = load_workbook(ruta_xlsx, data_only=True)
    facturas = defaultdict(lambda: {
        'fecha_factura': None,
        'total_factura': 0.0,
        'values': [],
        'source_file': archivo,
    })

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        if len(rows) < 2:
            logging.info(
                'Excel sin filas procesables | archivo=%s | hoja=%s',
                archivo,
                worksheet.title,
            )
            continue

        non_empty_row_numbers = [
            row_number
            for row_number, row in enumerate(rows[1:], start=2)
            if any(row)
        ]
        last_data_row = non_empty_row_numbers[-1] if non_empty_row_numbers else None

        try:
            columns = find_required_columns(rows[0])
        except util.PDFInconsistente as exc:
            raise ExcelRowError(str(exc), archivo, hoja=worksheet.title, fila=1) from exc

        for row_number, row in enumerate(rows[1:], start=2):
            if not any(row):
                logging.info(
                    'Fila vacia ignorada | archivo=%s | hoja=%s | fila=%s',
                    archivo,
                    worksheet.title,
                    row_number,
                )
                continue

            try:
                nro_factura = normalize_factura(row[columns['factura_oficial']])
                fecha_factura = format_excel_date(row[columns['fecha_factura']])
                nro_remito = extract_nro_remito(row[columns['remito']])
                tipo = str(row[columns['nombre_material']]).strip()
                total_pedido = to_float(row[columns['total_pedido']])
                certificados = parse_certificados(row[columns['numero_serie']])
            except util.PDFInconsistente as exc:
                if row_number == last_data_row:
                    logging.warning(
                        'Ultima fila inconsistente ignorada | archivo=%s | hoja=%s | fila=%s | detalle=%s',
                        archivo,
                        worksheet.title,
                        row_number,
                        str(exc),
                    )
                    continue
                raise ExcelRowError(
                    str(exc),
                    archivo,
                    hoja=worksheet.title,
                    fila=row_number,
                ) from exc

            cantidad_certificados = len(certificados)
            if cantidad_certificados == 0:
                if row_number == last_data_row:
                    logging.warning(
                        'Ultima fila inconsistente ignorada | archivo=%s | hoja=%s | fila=%s | detalle=%s',
                        archivo,
                        worksheet.title,
                        row_number,
                        'No se pudieron obtener certificados.',
                    )
                    continue
                raise ExcelRowError(
                    'No se pudieron obtener certificados.',
                    archivo,
                    hoja=worksheet.title,
                    fila=row_number,
                    factura=nro_factura,
                )

            precio_unitario = total_pedido / cantidad_certificados
            fecha_update = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            factura = facturas[nro_factura]

            if factura['fecha_factura'] is None:
                factura['fecha_factura'] = fecha_factura
            elif factura['fecha_factura'] != fecha_factura:
                if row_number == last_data_row:
                    logging.warning(
                        'Ultima fila inconsistente ignorada | archivo=%s | hoja=%s | fila=%s | detalle=%s',
                        archivo,
                        worksheet.title,
                        row_number,
                        'La factura tiene mas de una fecha en el workbook.',
                    )
                    continue
                raise ExcelRowError(
                    'La factura tiene más de una fecha en el workbook.',
                    archivo,
                    hoja=worksheet.title,
                    fila=row_number,
                    factura=nro_factura,
                )

            factura['total_factura'] += total_pedido
            logging.info(
                'Fila procesada | archivo=%s | hoja=%s | fila=%s | factura=%s | remito=%s | certificados=%s | total_pedido=%.2f | precio_unitario=%.2f | tipo=%s',
                archivo,
                worksheet.title,
                row_number,
                nro_factura,
                nro_remito,
                certificados,
                total_pedido,
                precio_unitario,
                tipo,
            )

            for certificado in certificados:
                factura['values'].append([
                    nro_remito,
                    certificado,
                    fecha_factura,
                    nro_factura,
                    precio_unitario,
                    fecha_update,
                    'proceso de facturación',
                    tipo,
                ])

    return facturas


def buscar_xlsx(directorio):
    if not os.path.isdir(directorio):
        return []

    archivos = []
    for archivo in os.listdir(directorio):
        if archivo.lower().endswith('.xlsx'):
            archivos.append(os.path.join(directorio, archivo))
    return sorted(archivos)


def mover_archivo(nombre_archivo, directorio_destino):
    if not os.path.isfile(nombre_archivo):
        raise FileNotFoundError(f"El archivo '{nombre_archivo}' no existe.")

    if not os.path.exists(directorio_destino):
        os.makedirs(directorio_destino)

    nombre_archivo_nuevo = os.path.join(directorio_destino, os.path.basename(nombre_archivo))
    shutil.move(nombre_archivo, nombre_archivo_nuevo)
    return nombre_archivo_nuevo


def build_html_inconsistencia(
    nro_factura,
    archivo,
    detalle,
    remitos_no_encontrados=None,
    certificados_no_encontrados=None,
    total_factura=None,
    total_calculado=None,
):
    remitos_html = ''
    if remitos_no_encontrados:
        remitos_html = f"""
        <p>Los siguientes Remitos no fueron encontrados en la base de datos de certificados de origen:</p>
        <ul>
            {''.join([f'<li>{remito}</li>' for remito in remitos_no_encontrados])}
        </ul>
        """

    certificados_html = ''
    if certificados_no_encontrados:
        certificados_html = f"""
        <p>Los siguientes certificados no fueron encontrados en la base de datos de certificados de origen:</p>
        <ul>
            {''.join([f"<li>Remito {item['nro_remito']} | Certificado {item['certificado'] or '(vacio)'}</li>" for item in certificados_no_encontrados])}
        </ul>
        """

    total_html = ''
    if total_factura is not None and total_calculado is not None:
        total_html = f"""
        <p>El total de la factura no coincide con el total calculado:</p>
        <ul>
            <li>Total factura: {total_factura}</li>
            <li>Total calculado: {total_calculado}</li>
        </ul>
        """

    return f"""<html>
    <body>
        <p>Para la factura: <strong>{nro_factura}</strong></p>
        <p>del archivo: {archivo}</p>
        <p>{detalle}</p>
        {remitos_html}
        {certificados_html}
        {total_html}
        <p>Por favor reenviar los datos a la casilla <strong>{os.getenv('EMAIL_USER', '')}</strong> para su procesamiento</p>
    </body>
    </html>"""


def normalize_mail_recipients(value):
    if isinstance(value, list):
        recipients = value
    else:
        recipients = re.split(r'[;,]', str(value or ''))

    return [recipient.strip() for recipient in recipients if recipient and recipient.strip()]


def html_to_plain_text(value):
    if not value:
        return ''

    if not value.lstrip().startswith('<'):
        return value

    text = re.sub(r'<li>\s*', '- ', value, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '\n', text)
    text = re.sub(r'[ \t]+\n', '\n', text)
    text = re.sub(r'\n{2,}', '\n', text)
    return text.strip()


def safe_send_mail(to, subject, plain_message, html_message):
    recipients = normalize_mail_recipients(to)
    fallback = normalize_mail_recipients(os.getenv('EMAIL_TICKETS', ''))
    if not recipients and fallback:
        recipients = fallback
        logging.warning(
            'EMAIL_ADMINISTRACION vacio. Se usa EMAIL_TICKETS como fallback | subject=%s',
            subject,
        )

    if not recipients:
        print(f'No hay destinatarios configurados para enviar el correo: {subject}')
        logging.error('No hay destinatarios configurados para enviar el correo | subject=%s', subject)
        return False

    if html_message and (not plain_message or plain_message.lstrip().startswith('<')):
        plain_message = html_to_plain_text(html_message)

    try:
        smtp.smtp.SendMail(recipients, subject, plain_message, html_message, '')
        logging.info('Correo enviado | to=%s | subject=%s', recipients, subject)
        return True
    except Exception:
        print(f'Fallo al enviar email | to={recipients} | subject={subject}')
        logging.exception('Fallo al enviar email | to=%s | subject=%s', recipients, subject)
        return False


def procesar_factura(dbase, archivo, nro_factura, data_factura):
    total_factura = round(data_factura['total_factura'], 2)
    total_calculado, remitos_no_encontrados, certificados_no_encontrados = dbase.CertificadoFactura(
        data_factura['values'],
        total_factura
    )
    total_calculado = round(total_calculado, 2)

    print(f'Factura: {nro_factura}')
    print(f'Total factura: {total_factura}')
    print(f'Total calculado: {total_calculado}')
    print(f'Remitos no encontrados: {remitos_no_encontrados}')
    print(f'Certificados no encontrados: {certificados_no_encontrados}')
    logging.info(
        'Factura procesada | archivo=%s | factura=%s | items=%s | total_factura=%.2f | total_calculado=%.2f | remitos_no_encontrados=%s | certificados_no_encontrados=%s',
        archivo,
        nro_factura,
        len(data_factura['values']),
        total_factura,
        total_calculado,
        remitos_no_encontrados,
        certificados_no_encontrados,
    )

    dbase.CertificadoDocumentoInsertOrUpdate([
        nro_factura,
        nro_factura,
        total_factura,
        total_calculado,
    ])

    if remitos_no_encontrados or certificados_no_encontrados:
        raise util.PDFInconsistente(
            build_html_inconsistencia(
                nro_factura,
                archivo,
                'Se encontraron remitos o certificados que no existen en la tabla de certificados.',
                remitos_no_encontrados=remitos_no_encontrados,
                certificados_no_encontrados=certificados_no_encontrados,
            )
        )

    if total_factura != total_calculado:
        raise util.PDFInconsistente(
            build_html_inconsistencia(
                nro_factura,
                archivo,
                'El total de la factura no coincide con el total calculado.',
                total_factura=total_factura,
                total_calculado=total_calculado,
            )
        )


def main():
    archivo = ''
    try:
        start_time = util.show_time('Inicio')
        attachments_folder = os.getenv('ATTACHMENTS_FOLDER', '')
        archivos_xlsx = buscar_xlsx(attachments_folder)
        dbase = db.DB()

        for archivo in archivos_xlsx:
            try:
                print(f'Archivo: {archivo}')
                logging.info('Inicio archivo Excel | archivo=%s', os.path.basename(archivo))
                facturas = extraer_facturas_desde_xlsx(archivo)
                if not facturas:
                    raise util.PDFInconsistente('El archivo no contiene filas procesables.')

                for nro_factura, data_factura in facturas.items():
                    procesar_factura(dbase, os.path.basename(archivo), nro_factura, data_factura)

                mover_archivo(archivo, os.getenv('PROCESSED_FOLDER', ''))
                logging.info('Archivo movido a procesados | archivo=%s', os.path.basename(archivo))
            except ExcelRowError as e:
                detalle = e.describe()
                print(f'Error: {detalle}')
                logging.error('Error de fila Excel | %s', detalle)
                html_msg_admin = build_html_inconsistencia(
                    e.factura or 'sin factura',
                    e.archivo,
                    detalle,
                )
                safe_send_mail(
                    os.getenv('EMAIL_ADMINISTRACION', ''),
                    f'Inconsistencia en Excel de Factura CAC ce Cert Origen {os.path.basename(archivo)}',
                    detalle,
                    html_msg_admin,
                )
            except util.PDFInconsistente as e:
                detalle = str(e)
                print(f'Error: {detalle}')
                logging.error('Error de negocio | archivo=%s | detalle=%s', os.path.basename(archivo), detalle)
                safe_send_mail(
                    os.getenv('EMAIL_ADMINISTRACION', ''),
                    f'Inconsistencia en procesar Factura CAC ce Cert Origen {os.path.basename(archivo)}',
                    detalle,
                    detalle if detalle.lstrip().startswith('<html>') else build_html_inconsistencia(
                        'sin factura',
                        os.path.basename(archivo),
                        detalle,
                    ),
                )
            except Exception:
                description_archivo = traceback.format_exc()
                logging.exception('Error no controlado procesando archivo | archivo=%s', os.path.basename(archivo))
                html_msg_admin = build_html_inconsistencia(
                    'sin factura',
                    os.path.basename(archivo),
                    description_archivo,
                )
                safe_send_mail(
                    os.getenv('EMAIL_ADMINISTRACION', ''),
                    f'Error no controlado procesando Factura CAC ce Cert Origen {os.path.basename(archivo)}',
                    description_archivo,
                    html_msg_admin,
                )

        end_time = util.show_time('Fin')
        print(f'Tiempo de ejecución: {end_time - start_time}')

    except Exception as e:
        description = traceback.format_exc()
        traceback.print_exc()
        print(description)
        frame = inspect.currentframe()
        function_name = inspect.getframeinfo(frame).function
        print(f'Error ocurrió en la función: {function_name}')
        print(f'Archivo: {inspect.getfile(frame)}')
        print(f'Error: {e}')

        html_msg = f"""<html>
        <body>
            <p>Error ocurrido en la función: {function_name}</p>
            <p>Archivo: {inspect.getfile(frame)}</p>
            <p>Error: {e}</p>
            <p>Descripción: {description}</p>
        </body>
        </html>"""
        logging.exception('Error fatal en extraer_factura.main')
        safe_send_mail(
            os.getenv('EMAIL_TICKETS', ''),
            f'Lectura de Factura {os.path.basename(archivo)}',
            f'{description} {frame} {function_name}',
            html_msg,
        )


if __name__ == '__main__':
    main()
