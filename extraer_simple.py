import traceback
import fitz  # PyMuPDF
import re
from dotenv import load_dotenv
import os
import logging
import fnmatch
import shutil
import pandas as pd
from openpyxl import load_workbook
load_dotenv()

log_level = os.getenv('LOG_LEVEL', 'INFO')  # Valor por defecto 'INFO' si no está definido

# Configurar el nivel de logging dinámicamente
numeric_level = getattr(logging, log_level.upper(), logging.INFO)

logging.basicConfig(
    level=numeric_level,  # Nivel de registro mínimo que quieres mostrar
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
def expandir_rango(rango_str):
    if "al" in rango_str:
        partes = rango_str.split("al")
        inicio = int(partes[0].strip())
        fin = int(partes[1].strip())
        return [str(num).zfill(len(partes[0].strip())) for num in range(inicio, fin + 1)]
    return [rango_str.strip()]

def extraer_informacion_pdf(ruta_pdf):
    # Abrir el archivo PDF
    documento = fitz.open(ruta_pdf)
    texto_completo = ""
    for pagina in documento:
        texto_completo += pagina.get_text()

    # Expresiones regulares para buscar la información
    

    coordenadas_area_fecha = (310, 42, 510, 80) # Ajusta las coordenadas de acuerdo a tus necesidades
    x1, y1, x2, y2 = coordenadas_area_fecha
    area_fecha = fitz.Rect(x1, y1, x2, y2)  
    texto_area = pagina.get_textbox(area_fecha)

    rx_valor = re.search(r'RX0001\s*(\d+)', texto_area)
    rx_encontrado = rx_valor.group(1) if rx_valor else "No encontrado"
    fecha = re.search(r'\d{1,2}/\d{1,2}/\d{4}', texto_area)
    fechas = re.findall(r'\d{1,2}/\d{1,2}/\d{4}', texto_completo)
    fecha_encontrada = fecha.group() if fecha else "No encontrado"

    certificados = re.search(r'Certificado:\s*([\d\s\w]+)', texto_completo)
    certificados_encontrados = certificados.group(1).strip() if certificados else "No encontrado"

    return rx_encontrado, fecha_encontrada, certificados_encontrados


def buscar_pdfs_con_nombre_similar(directorio, patron="CAC01005448-ORG-RX0001-*.pdf"):
    """
    Busca archivos PDF en un directorio cuyo nombre coincida con un patrón dado.

    Parámetros:
    - directorio (str): Ruta del directorio donde buscar los archivos.
    - patron (str): Patrón de búsqueda para coincidir con los nombres de archivo (por defecto busca nombres similares a "CAC01005448-ORG-RX0001-*.pdf").

    Retorna:
    - List[str]: Una lista de rutas completas a los archivos que coinciden con el patrón.
    """
    archivos_pdf = []
    # Recorre los archivos en el directorio especificado
    for archivo in os.listdir(directorio):
        if fnmatch.fnmatch(archivo, patron):
            # Agrega la ruta completa del archivo a la lista
            archivos_pdf.append(os.path.join(directorio, archivo))
    return archivos_pdf


def mover_archivo(nombre_archivo, directorio_destino):
    """
    Mueve un archivo a un directorio específico.

    Parámetros:
    - nombre_archivo (str): Ruta completa del archivo que deseas mover.
    - directorio_destino (str): Ruta del directorio destino.

    Retorna:
    - str: Ruta completa del archivo movido en el nuevo directorio.
    """
    # Verifica si el archivo existe
    if not os.path.isfile(nombre_archivo):
        raise FileNotFoundError(f"El archivo '{nombre_archivo}' no existe.")
    
    # Verifica si el directorio destino existe, si no, lo crea
    if not os.path.exists(directorio_destino):
        os.makedirs(directorio_destino)

    # Construye la ruta del nuevo archivo
    nombre_archivo_nuevo = os.path.join(directorio_destino, os.path.basename(nombre_archivo))

    # Mueve el archivo
    shutil.move(nombre_archivo, nombre_archivo_nuevo)

    return nombre_archivo_nuevo

def agregar_filas_sin_duplicados_excel(archivo_excel, nuevas_filas):
    """
    Agrega nuevas filas al final de un archivo Excel si no existen.

    Parámetros:
    - archivo_excel (str): Ruta del archivo Excel existente.
    - nuevas_filas (list of dict): Lista de nuevas filas a agregar. Cada fila se representa como un diccionario con clave-valor correspondiente a las columnas.

    Retorna:
    - None
    """
    # Verifica si el archivo existe
    if os.path.isfile(archivo_excel):
        # Carga el archivo existente
        df_existente = pd.read_excel(archivo_excel)
    else:
        # Si el archivo no existe, crea un DataFrame vacío
        df_existente = pd.DataFrame()

    # Convierte las nuevas filas a un DataFrame
    df_nuevas_filas = pd.DataFrame(nuevas_filas)

    #if 'Fecha Remito' in df_nuevas_filas.columns:
    #    df_nuevas_filas['Fecha Remito'] = pd.to_datetime(df_nuevas_filas['Fecha Remito'], errors='coerce').dt.strftime('%d/%m/%Y')  # Convertir a fecha
    if 'Certificado' in df_nuevas_filas.columns:
        df_nuevas_filas['Certificado'] = df_nuevas_filas['Certificado'].astype(str)  # Convertir a string

    # Verifica si hay duplicados y elimina las filas que ya existen
    if not df_existente.empty:
        # Compara las filas y descarta las que ya existen en el archivo
        df_actualizado = pd.concat([df_existente, df_nuevas_filas], ignore_index=True).drop_duplicates(keep='first')
    else:
        # Si el archivo está vacío, simplemente usa las nuevas filas
        df_actualizado = df_nuevas_filas

    # Guarda el archivo actualizado
    df_actualizado.to_excel(archivo_excel, index=False)
def verificar_y_agregar_filas_excel(archivo_excel, nuevas_filas, columnas, hoja='Sheet1'):
    """
    Verifica si las filas ya existen en un archivo Excel y las agrega si no están presentes.

    Parámetros:
    - archivo_excel (str): Ruta del archivo Excel.
    - nuevas_filas (list of dict): Lista de nuevas filas a verificar y agregar.
    - columnas (list of str): Lista de nombres de las columnas para identificar duplicados.
    - hoja (str): Nombre de la hoja donde agregar las filas (por defecto 'Sheet1').

    Retorna:
    - None
    """
    # Carga el archivo Excel
    wb = load_workbook(archivo_excel)
    if hoja not in wb.sheetnames:
        raise ValueError(f"La hoja '{hoja}' no existe en el archivo Excel.")
    ws = wb[hoja]

    # Obtener encabezados
    headers = [cell.value for cell in ws[1]]
    # Crear conjunto para valores existentes
    valores_existentes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        valores = tuple(row[headers.index(col)] for col in columnas)
        valores_existentes.add(valores)

    # Agregar nuevas filas si no existen
    for fila in nuevas_filas:
        valores_nuevos = tuple(fila.get(col, "") for col in columnas)
        if valores_nuevos not in valores_existentes:
            nueva_fila = [fila.get(header, "") for header in headers]
            ws.append(nueva_fila)
            valores_existentes.add(valores_nuevos)

    # Guardar archivo
    wb.save(archivo_excel)
def main():
    try:
        attachments_folder = os.getenv('ATTACHMENTS_FOLDER', '') 
        # Rutas a los archivos PDF (ajusta según sea necesario)
        archivos_pdf = buscar_pdfs_con_nombre_similar(attachments_folder, patron="CAC01005448-ORG-RX0001-*.pdf")
        filas = []
        for archivo in archivos_pdf:
            rx, fecha, certificados = extraer_informacion_pdf(archivo)
            print(f"Archivo: {archivo}")
            print(f"RX0001: {rx}")
            print(f"Primera Fecha: {fecha}")
            print(f"Certificados: {expandir_rango(certificados)}")  # Expandir el rango de certificados y mostrarlos separados por comascertificados}")
            print("-" * 40)
            
            for certificado in expandir_rango(certificados):
                filas.append([{"Remito": rx, "Fecha Remito": fecha, "Certificado": certificado}])

            
            # Mover el archivo a la carpeta "procesados"
            mover_archivo(archivo, os.getenv('PROCESSED_FOLDER', ''))
        if filas != []:
            columnas_para_verificar = ['Remito', 'Fecha Remito', 'Certificado']
            verificar_y_agregar_filas_excel(os.getenv('EXCEL_FILE', ''), filas, columnas_para_verificar)

            #agregar_filas_sin_duplicados_excel(os.getenv('EXCEL_FILE', ''), filas)
            print(filas)
    except Exception as e:
        traceback.print_exc()
        print(f"Error: {e}")

main()